"""
Extracts the Smith-Wilson curve parameters (Coupon_freq, LLP, Convergence,
UFR, alpha, CRA, VA) per country from the
EIOPA_RFR_YYYYMMDD_Term_Structures.xlsx workbook.

Relevant sheets: 'RFR_spot_no_VA' and 'RFR_spot_with_VA'.
Layout (discovered by inspection):
  - Row 2: full country/area name per column
  - Row 3: an EIOPA code string like 'AT_30_11_2024_SWP_LLP_20_EXT_40_UFR_3.30'
           -- the leading token is the ISO-ish country code, and it also
           encodes the underlying instrument type (SWP/GOV/OIS/PEE).
  - Rows 4-9: Coupon_freq, LLP, Convergence, UFR, alpha, CRA (one value per country)
  - Row 10: VA (only populated with a real value on the *with_VA* sheet)

Output: one row per (reference_date, curve_type, country) with all
parameters as columns (wide format -- these are scalars, not a vector,
so wide is the natural shape; unlike the Qb vectors there's no repeating
index to pivot on).
"""
from __future__ import annotations

import openpyxl
import pandas as pd

SHEET_CURVE_TYPE = {
    "RFR_spot_no_VA": "no_VA",
    "RFR_spot_with_VA": "with_VA",
}

PARAM_ROW_LABELS = {
    4: "coupon_freq",
    5: "llp",
    6: "convergence",
    7: "ufr",
    8: "alpha",
    9: "cra",
    10: "va",
}

COUNTRY_CODE_ROW = 3
FIRST_DATA_COL = 3  # column C; column B holds row labels


def _normalise_country(country: str) -> str:
    """Align Term-Structures codes with Qb sheet labels used elsewhere."""
    if country == "EUR":
        return "EU"
    # EIOPA used GB_… through 2023-04, then UK_…; Qb sheets always label UK.
    if country == "GB":
        return "UK"
    return country


def _parse_country_code(code_string: str) -> tuple[str, str | None]:
    """
    Split 'AT_30_11_2024_SWP_LLP_20_EXT_40_UFR_3.30' into
    ('AT', 'SWP'). Also normalises EUR->EU and GB->UK so country codes
    match qb_vectors.csv / the notebook country list.
    """
    parts = code_string.split("_")
    # Expected shape: CODE_DD_MM_YYYY_INSTRUMENT_...
    if len(parts) <= 4:
        return _normalise_country(parts[0]), None

    country, _day, _month, _year, instrument, *_rest = parts
    return _normalise_country(country), instrument


def _extract_sheet(ws, curve_type: str, reference_date: str) -> pd.DataFrame:
    records = []

    for col in range(FIRST_DATA_COL, ws.max_column + 1):
        code_string = ws.cell(row=COUNTRY_CODE_ROW, column=col).value
        if not code_string:
            continue
        country, instrument_type = _parse_country_code(code_string)

        rec = {
            "reference_date": reference_date,
            "curve_type": curve_type,
            "country": country,
            "instrument_type": instrument_type,
        }
        for row, label in PARAM_ROW_LABELS.items():
            val = ws.cell(row=row, column=col).value
            rec[label] = float(val) if isinstance(val, (int, float)) else None

        records.append(rec)

    return pd.DataFrame.from_records(records)


def extract_parameters(xlsx_path: str, reference_date: str) -> pd.DataFrame:
    """
    Parse an EIOPA_RFR_YYYYMMDD_Term_Structures.xlsx file into a tidy
    DataFrame of curve parameters, one row per country per curve type.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    frames = []
    for sheet_name, curve_type in SHEET_CURVE_TYPE.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Expected sheet '{sheet_name}' not found in {xlsx_path}. "
                f"Found: {wb.sheetnames}"
            )
        frames.append(_extract_sheet(wb[sheet_name], curve_type, reference_date))

    return pd.concat(frames, ignore_index=True)


if __name__ == "__main__":
    import sys

    from config import extract_reference_date  # noqa: E402

    path = sys.argv[1]
    df = extract_parameters(path, extract_reference_date(path))
    print(df.head(20))
    print(f"\nTotal rows: {len(df)}")
    print(f"Countries: {df['country'].nunique()}")
