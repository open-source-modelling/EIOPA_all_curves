"""
Extracts the official EIOPA spot / yield curves from the
EIOPA_RFR_YYYYMMDD_Term_Structures.xlsx workbook.

Relevant sheets: 'RFR_spot_no_VA' and 'RFR_spot_with_VA' — the same sheets
used by param_extractor.py. After the fixed parameter block (rows 4-10),
each country column holds a spot-rate curve:

  - Column B: maturity in years (term_index), shared across countries
  - Columns C onward: spot rate (decimal, e.g. 0.02607 = 2.607 %)
  - Data runs from row 11 down to the first blank maturity cell
    (currently maturities 1..150)

Output: a tidy ("long") DataFrame with one row per
(reference_date, curve_type, country, term_index, spot_rate).
Country codes follow curve_parameters.csv (EUR -> EU).
"""
from __future__ import annotations

import openpyxl
import pandas as pd

from extractors.param_extractor import (
    COUNTRY_CODE_ROW,
    FIRST_DATA_COL,
    PARAM_ROW_LABELS,
    SHEET_CURVE_TYPE,
    _parse_country_code,
)

# Spot rates begin on the row immediately below the last parameter (VA).
FIRST_SPOT_ROW = max(PARAM_ROW_LABELS) + 1
TERM_INDEX_COL = 2  # column B


def _extract_sheet(ws, curve_type: str, reference_date: str) -> pd.DataFrame:
    # Country columns mirror param_extractor: code string on row 3, rates below.
    countries = {}
    for col in range(FIRST_DATA_COL, ws.max_column + 1):
        code_string = ws.cell(row=COUNTRY_CODE_ROW, column=col).value
        if not code_string:
            continue
        country, _instrument = _parse_country_code(code_string)
        countries[col] = country

    records = []
    for row in range(FIRST_SPOT_ROW, ws.max_row + 1):
        term_index = ws.cell(row=row, column=TERM_INDEX_COL).value
        if term_index is None:
            break  # end of the shared maturity column
        if not isinstance(term_index, (int, float)):
            continue

        for col, country in countries.items():
            spot_rate = ws.cell(row=row, column=col).value
            if spot_rate is None:
                continue
            if not isinstance(spot_rate, (int, float)):
                continue
            records.append(
                {
                    "reference_date": reference_date,
                    "curve_type": curve_type,
                    "country": country,
                    "term_index": float(term_index),
                    "spot_rate": float(spot_rate),
                }
            )

    return pd.DataFrame.from_records(records)


def extract_yield_curves(xlsx_path: str, reference_date: str) -> pd.DataFrame:
    """
    Parse an EIOPA_RFR_YYYYMMDD_Term_Structures.xlsx file into a tidy
    long-format DataFrame of official spot rates for both curve types.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    frames = []
    for sheet_name, curve_type in SHEET_CURVE_TYPE.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Expected sheet '{sheet_name}' not found in {xlsx_path}. "
                f"Found: {wb.sheetnames}"
            )
        frames.append(_extract_sheet(wb[sheet_name], curve_type, reference_date))

    return pd.concat(frames, ignore_index=True)


if __name__ == "__main__":
    import sys

    from config import extract_reference_date  # noqa: E402

    path = sys.argv[1]
    df = extract_yield_curves(path, extract_reference_date(path))
    print(df.head(20))
    print(f"\nTotal rows: {len(df)}")
    print(f"Countries: {df['country'].nunique()}")
    print(f"Curve types: {df['curve_type'].unique().tolist()}")
    print(f"Term range: {df['term_index'].min()} .. {df['term_index'].max()}")
