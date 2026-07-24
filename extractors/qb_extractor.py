"""
Extracts the Smith-Wilson Qb calibration vectors from the
EIOPA_RFR_YYYYMMDD_Qb_SW.xlsx workbook.

Layout of each sheet (SW_Qb_no_VA / SW_Qb_with_VA), discovered by inspection:
  - Row 4: title "PARAMETERS OF THE SMITH-WILSON EXTRAPOLATION METHOD"
  - Header row (country codes) is the first row where more than one cell
    contains a short (<=3 char) text label -- for the Nov-2024 file this is
    row 10, but we detect it rather than hard-coding the row number, since
    EIOPA occasionally shifts layout by a row or two between releases.
  - Country blocks repeat every 3 columns: [index, Qb value, blank spacer].
  - Data rows run from the row right below the header down to the first
    fully-blank row.

Output: a tidy ("long") DataFrame with one row per
(reference_date, curve_type, country, term_index, qb_value).
"""
from __future__ import annotations

import openpyxl
import pandas as pd

SHEET_CURVE_TYPE = {
    "SW_Qb_no_VA": "no_VA",
    "SW_Qb_with_VA": "with_VA",
}

# Header-detection heuristics: a real country-header row has dozens of short
# ISO-ish codes (AT, BE, EUR, ...), not titles or numeric cells.
MAX_COUNTRY_CODE_LEN = 4
MIN_HEADER_COUNTRY_LABELS = 5
MAX_HEADER_SCAN_ROWS = 25

# Each country occupies three columns: [term_index, qb_value, blank spacer].
COUNTRY_BLOCK_WIDTH = 3


def _find_header_row(ws, max_scan_rows: int = MAX_HEADER_SCAN_ROWS) -> int:
    """Find the row containing the country-code labels (e.g. 'EUR', 'AT', 'BE')."""
    for row in range(1, max_scan_rows + 1):
        labels = [
            ws.cell(row=row, column=col).value
            for col in range(1, ws.max_column + 1)
        ]
        text_labels = [
            v for v in labels
            if isinstance(v, str) and 1 <= len(v) <= MAX_COUNTRY_CODE_LEN
        ]
        if len(text_labels) >= MIN_HEADER_COUNTRY_LABELS:
            return row
    raise ValueError(f"Could not locate the country header row in sheet '{ws.title}'")


def _extract_sheet(ws, curve_type: str, reference_date: str) -> pd.DataFrame:
    """
    Walk one Qb sheet into a long DataFrame.

    Column geometry (1-based Excel columns), repeating every COUNTRY_BLOCK_WIDTH::

        | term_index | qb_value | spacer | term_index | qb_value | spacer | ...
        |     1      |   AT     |        |     4      |   BE     |        | ...
        |    1.0     |  0.12    |        |    1.0     |  0.09    |        | ...

    Only header cells with a string label are countries; the unlabeled columns
    are either the term-index column (one left of the value) or the blank spacer.
    The Qb sheet labels the euro-area block 'EUR'; we normalise that to 'EU'
    so it matches curve_parameters.csv / yield_curves.csv.
    """
    header_row = _find_header_row(ws)

    # Value column = middle of each 3-column block (the cell that holds the code).
    country_cols = {
        col: ws.cell(row=header_row, column=col).value
        for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=header_row, column=col).value, str)
    }

    records = []
    for value_col, country in country_cols.items():
        # Align with param/spot extractors: EIOPA Qb header says EUR, we store EU.
        if country == "EUR":
            country = "EU"
        index_col = value_col - 1  # term index sits one column left of qb_value
        for row in range(header_row + 1, ws.max_row + 1):
            term_index = ws.cell(row=row, column=index_col).value
            qb_value = ws.cell(row=row, column=value_col).value
            if term_index is None and qb_value is None:
                break  # end of this country block
            if term_index is None or qb_value is None:
                continue  # partial row: skip
            records.append(
                {
                    "reference_date": reference_date,
                    "curve_type": curve_type,
                    "country": country,
                    # Kept as float, not int: most countries index by whole
                    # years (1, 2, 3 ...) but some (e.g. AU, CA, CN, HK, MX,
                    # NZ, SG, ZA, KR, TH) use semi-annual steps (0.5, 1, 1.5 ...)
                    # reflecting their coupon frequency. Truncating to int would
                    # collide 0.5/1.0 and 1.5/2.0 into the same key.
                    "term_index": float(term_index),
                    "qb_value": float(qb_value),
                }
            )

    return pd.DataFrame.from_records(records)


def extract_qb_vectors(xlsx_path: str, reference_date: str) -> pd.DataFrame:
    """
    Parse an EIOPA_RFR_YYYYMMDD_Qb_SW.xlsx file into a tidy long-format
    DataFrame covering both the no-VA and with-VA curves, for every country.
    """
    # Note: read_only=True is NOT used here. openpyxl's read-only mode is
    # optimized for sequential row iteration; the random row/column .cell()
    # access this parser relies on is quadratic in that mode (50s+ on this
    # file) versus well under a second in normal mode.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    frames = []
    for sheet_name, curve_type in SHEET_CURVE_TYPE.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Expected sheet '{sheet_name}' not found in {xlsx_path}. "
                f"Found: {wb.sheetnames}"
            )
        frames.append(_extract_sheet(wb[sheet_name], curve_type, reference_date))

    return pd.concat(frames, ignore_index=True)


if __name__ == "__main__":
    import sys

    from config import extract_reference_date  # noqa: E402

    path = sys.argv[1]
    df = extract_qb_vectors(path, extract_reference_date(path))
    print(df.head(20))
    print(f"\nTotal rows: {len(df)}")
    print(f"Countries: {df['country'].nunique()}")
    print(f"Curve types: {df['curve_type'].unique().tolist()}")
